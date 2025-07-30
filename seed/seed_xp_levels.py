import os
from openpyxl import load_workbook
from tactera_backend.core.database import get_session
from tactera_backend.models.stat_level_requirement import StatLevelRequirement

from sqlalchemy import inspect
from tactera_backend.core.database import get_session
from tactera_backend.models.stat_level_requirement import StatLevelRequirement

def debug_print_stat_table():
    session = next(get_session())
    insp = inspect(session.bind)
    cols = insp.get_columns("stat_level_requirement")
    print("\n=== DEBUG: Database Columns in stat_level_requirement ===")
    for col in cols:
        print(f"{col['name']} (nullable={col['nullable']})")

    print("\n=== DEBUG: Model Columns in StatLevelRequirement ===")
    print(StatLevelRequirement.__table__.columns.keys())
# END


def safe_seed_stat_levels():
    """
    Seeds the StatLevelRequirement table from an Excel file (xp_levels.xlsx).
    Only runs if the table is currently empty.
    """
    session = next(get_session())

    existing_count = session.query(StatLevelRequirement).count()
    if existing_count > 0:
        print(f"⚠️ StatLevelRequirement already has {existing_count} rows. Skipping seeding.")
        return

    try:
        # Determine the path to xp_levels.xlsx relative to this script
        current_file = os.path.abspath(__file__)
        current_dir = os.path.dirname(current_file)  # seed/
        project_root = os.path.dirname(current_dir)  # tactera_backend/
        excel_path = os.path.join(project_root, "xp_levels.xlsx")

        wb = load_workbook(excel_path)
        sheet = wb.active

        inserted = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            level, xp = row
            if level is None or xp is None:
                continue
            session.add(StatLevelRequirement(level=int(level), xp_required=int(xp)))
            inserted += 1
            
            
            for obj in session.new:
                print("\n=== DEBUG: Object Pending Insert ===")
                print(obj)
                print(obj.__dict__)  # Shows all fields tracked by SQLAlchemy


        session.commit()
        print(f"✅ Seeded {inserted} levels into StatLevelRequirement.")

    except FileNotFoundError:
        print(f"❌ xp_levels.xlsx not found at expected path: {excel_path}")
    except Exception as e:
        print(f"❌ Seeding failed: {e}")
