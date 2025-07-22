from openpyxl import load_workbook
from models import StatLevelRequirement
from database import get_sync_session



from openpyxl import load_workbook
from models import StatLevelRequirement
from database import get_session  # Uses the same session method your API routes use

def safe_seed_stat_levels():
    """
    Seeds the StatLevelRequirement table from an Excel file (xp_levels.xlsx).
    Only runs if the table is currently empty.
    """
    # Get a database session
    session = next(get_session())

    # Check if data already exists in the table
    existing_count = session.query(StatLevelRequirement).count()
    if existing_count > 0:
        print(f"⚠️ StatLevelRequirement already has {existing_count} rows. Skipping seeding.")
        return

    try:
        # Load the Excel workbook and get the active sheet
        wb = load_workbook("xp_levels.xlsx")
        sheet = wb.active

        inserted = 0  # Keep track of how many levels we insert

        # Loop through each row, starting from row 2 to skip headers
        for row in sheet.iter_rows(min_row=2, values_only=True):
            level, xp = row

            # Skip rows with missing values
            if level is None or xp is None:
                continue

            # Add a new row to the database
            session.add(StatLevelRequirement(level=int(level), xp_required=int(xp)))
            inserted += 1

        # Save all changes to the database
        session.commit()
        print(f"✅ Seeded {inserted} levels into StatLevelRequirement.")

    except FileNotFoundError:
        print("❌ xp_levels.xlsx not found. Make sure it's in the same folder as main.py.")
    except Exception as e:
        print(f"❌ Seeding failed: {e}")
